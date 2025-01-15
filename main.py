from __future__ import annotations

import shutil
import sys

import pandas as pd
import os
import zipfile
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

logger = logging.getLogger(__name__)
LOGGING_LEVEL = logging.DEBUG
logger.setLevel(LOGGING_LEVEL)

FORMAT_CURRENCY_RUB = "0.00 ₽"


def get_script_folder():
    # path of main .py or .exe when converted with pyinstaller
    # путь к файлу main .py или .exe при преобразовании с помощью pyinstaller
    if getattr(sys, 'frozen', False):
        script_path = os.path.dirname(sys.executable)
    else:
        script_path = os.path.dirname(
            os.path.abspath(sys.modules['__main__'].__file__)
        )
    return script_path


def transform_string(input_string: str | int):
    if isinstance(input_string, int):
        input_string = str(input_string)
    # Step 1: Remove the quotation marks
    # Убираем кавычки
    cleaned_string = input_string.replace('«', '').replace('»', '')
    cleaned_string = cleaned_string.replace('"', '')

    # Step 2: Trim any unwanted characters (if necessary)
    # Удалите все ненужные символы (при необходимости)
    # In this case, no additional trimming is needed, but this step is included for completeness
    # В этом случае дополнительная обрезка не требуется, но этот шаг включен для полноты картины
    return cleaned_string


def read_excel_file(file_path, sheet: str):
    # Считывает файл Excel и возвращает фрейм данных
    """
    Reads an Excel file and returns a DataFrame.
    """
    return pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet)


def write_to_excel(df: pd.DataFrame, company_name: str, company_id: str, file_path: str) -> None:
    # Функция получает фрейм данных и записывает его в xlsx-файл в соответствии с шаблоном
    """
    Function receives dataframe and writes into a xlsx file according to template
    """
    current_section_top_row: int = 1  # the top row of the current section
    # Add two empty columns to the DataFrame
    # Добавляем два пустых столбца во фрейм данных
    df = df.assign(EmptyColumn1=None, EmptyColumn2=None)

    # Convert the filtered DataFrame to a list of tuples
    # Преобразуем отфильтрованный фрейм данных в список кортежей
    data_to_write = df.values.tolist()

    # Define the headers
    # Определяем заголовки таблицы
    headers = ['№ накладной', 'Дата отпр.', 'Город отправителя', 'Город получателя', 'Вес',
               "Объем.вес", "Страховка", "Стоимость"]

    # Create a new workbook and select the active worksheet
    # Создаем новую рабочую книгу и выбираем активный рабочий лист
    wb = Workbook()
    ws = wb.active

    # Save the first part - the name of the table
    # Сохраняем первую часть - название таблицы
    # Merge cells from A2 to J2 and add text
    # Объединяем ячейки от A2 до J2 и добавляем текст
    current_section_top_row += 1
    ws.merge_cells(f'A{current_section_top_row}:J{current_section_top_row}')
    merged_cell = ws[f'A{current_section_top_row}']
    merged_cell.value = "Детализация реестра отправлений корреспонденции и грузов"
    merged_cell.font = Font(name='Arial Cyr', size=15)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Define border style
    # Определение стиля границы
    thin_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'),
                         bottom=Side(border_style='thin'))
    thick_left_border = Border(left=Side(border_style='thick'), right=Side(border_style='thin'),
                               top=Side(border_style='thin'), bottom=Side(border_style='thin'))
    thick_top_left_border = Border(left=Side(border_style='thick'), right=Side(border_style='thin'),
                                   top=Side(border_style='thick'), bottom=Side(border_style='thin'))
    thick_top_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                              top=Side(border_style='thick'), bottom=Side(border_style='thin'))
    thick_top_right_border = Border(left=Side(border_style='thin'), right=Side(border_style='thick'),
                                    top=Side(border_style='thick'), bottom=Side(border_style='thin'))
    thick_right_border = Border(left=Side(border_style='thin'), right=Side(border_style='thick'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
    thick_bottom_left_border = Border(left=Side(border_style='thick'), right=Side(border_style='thin'),
                                      top=Side(border_style='thin'), bottom=Side(border_style='thick'))
    thick_bottom_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                 top=Side(border_style='thin'), bottom=Side(border_style='thick'))
    thick_bottom_right_border = Border(left=Side(border_style='thin'), right=Side(border_style='thick'),
                                       top=Side(border_style='thin'), bottom=Side(border_style='thick'))

    # ---------------------------------------------------------
    # Save the second part - the name and the id of the company
    # Сохраняем вторую часть - название и идентификатор компании
    # ---------------------------------------------------------
    current_section_top_row += 2
    ws.merge_cells(f'A{current_section_top_row}:D{current_section_top_row}')
    merged_cell = ws[f'A{current_section_top_row}']
    merged_cell.value = company_name
    merged_cell.font = Font(name='Arial Cyr', size=11)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws[f'F{current_section_top_row}'].value = company_id
    ws[f'F{current_section_top_row}'].font = Font(name='Arial Cyr', size=11)
    ws[f'F{current_section_top_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # ---------------------------------------------------------
    # Save the third part - the imported data
    # Сохраняем третью часть - импортированные данные
    # ---------------------------------------------------------
    current_section_top_row += 2
    # Write the headers to the worksheet
    # Записываем заголовки на рабочий лист
    # Append the headers to the newly inserted row
    # Добавляем заголовки к только что вставленной строке

    thick_border = Border(left=Side(border_style='thick'), right=Side(border_style='thick'),
                          top=Side(border_style='thick'), bottom=Side(border_style='thick'))
    for header in headers:
        ws.cell(row=current_section_top_row, column=headers.index(header) + 1, value=header).border = thick_border

    current_section_top_row += 1

    # Write the data to the Excel file, starting from row 6 to avoid overwriting the headers
    # Записываем данные в файл Excel, начиная со строки 6, чтобы избежать перезаписи заголовков
    for i, row in enumerate(data_to_write, start=current_section_top_row):
        print(
            f"Row {i}: {row}")  # Adjust the print statement as needed. Отрегулируйте инструкцию по печати по мере необходимости
        _cell = ws.cell(row=i, column=8)
        _cell.number_format = FORMAT_CURRENCY_RUB
        for j, cell in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=cell)
            # Apply border to the first and last row, and first and last column
            # Примените границу к первой и последней строке, а также к первому и последнему столбцу
            # left top corner is i == 6 | j == 1
            # левый верхний угол равен i == 6 | j == 1
            if i == current_section_top_row and j == 1:
                ws.cell(row=i, column=j).border = thick_top_left_border
            # right top corner is i == 6 | j == len(row) + 1
            # правый верхний угол равен i == 6 | j == длина(строка) + 1
            elif i == current_section_top_row and j == len(row):
                ws.cell(row=i, column=j).border = thick_top_right_border
            # left bottom corner
            # левый нижний угол
            elif i == len(data_to_write) + current_section_top_row - 1 and j == 1:
                ws.cell(row=i, column=j).border = thick_bottom_left_border
            # right bottom corner
            # правый нижний угол
            elif i == len(data_to_write) + current_section_top_row - 1 and j == len(row):
                ws.cell(row=i, column=j).border = thick_bottom_right_border
            # top row
            # верхний ряд
            elif i == current_section_top_row:
                ws.cell(row=i, column=j).border = thick_top_border
            # bottom row
            # нижний ряд
            elif i == len(data_to_write) + current_section_top_row - 1:
                ws.cell(row=i, column=j).border = thick_bottom_border
            # right column
            # правая колонка
            elif j == len(row):
                ws.cell(row=i, column=j).border = thick_right_border
            # left column
            # левая колонка
            elif j == 1:
                ws.cell(row=i, column=j).border = thick_left_border
            else:
                ws.cell(row=i, column=j).border = thin_border

    # Define the range of rows you're interested in
    # Определите диапазон интересующих вас строк
    start_row = current_section_top_row - 1
    end_row = 41

    # Iterate through each column in the worksheet
    # Выполните итерацию по каждому столбцу на рабочем листе
    for column in ws.columns:
        # Initialize max_width to 0 for each column
        # Инициализируйте значение max_width равным 0 для каждого столбца
        max_width = 0
        # Iterate through the specified range of rows for the current column
        # Выполните итерацию по указанному диапазону строк для текущего столбца
        for cell in column:
            # Check if the cell is within the specified range
            # Проверьте, находится ли ячейка в пределах указанного диапазона
            if start_row <= cell.row <= end_row:
                # Find the maximum width of the cells in the column within the specified range
                # Найдите максимальную ширину ячеек в столбце в пределах указанного диапазона
                max_width = max(max_width, len(str(cell.value)))
        # Set the width of the column to the maximum width
        # Установите ширину столбца на максимальную ширину
        ws.column_dimensions[column[0].column_letter].width = max_width + 3

    # Set the width of column "F"
    # Установите ширину столбца "F"
    ws.column_dimensions['F'].width = max(ws.column_dimensions['F'].width, len(str(ws['F4'].value))) + 3

    # Save the fourth part - totals data
    # Сохраните четвертую часть - итоговые данные
    current_section_top_row += len(data_to_write) + 4
    # Totals Row #1 - Merge cells from D to F and add text
    # Итоговая строка #1 - Объединяем ячейки от D до F и добавляем текст
    ws.merge_cells(f'D{current_section_top_row}:F{current_section_top_row}')
    merged_cell = ws[f'D{current_section_top_row}']
    merged_cell.value = "Курьерские услуги:"
    merged_cell.font = Font(name='Arial Cyr', size=11)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thick_top_left_border

    # merged cell borders in column E and F are not shown in Microsoft Office Excel
    # границы объединенных ячеек в столбцах E и F не отображаются в Microsoft Office Excel
    # but do show in LibreOffice Calc, I think that it happens because of a merge cell coordinates
    # но показываются в LibreOffice Calc, я думаю, что это происходит из-за слияния координат ячеек
    # so border should be drawn explicitly in that columns
    # таким образом, граница должна быть явно прорисована в этих столбцах

    ws[f'E{current_section_top_row}'].border = thick_top_border
    ws[f'F{current_section_top_row}'].border = thick_top_border

    ws[f'G{current_section_top_row}'].value = f""
    ws[f'G{current_section_top_row}'].font = Font(name='Arial Cyr', size=9)
    ws[f'G{current_section_top_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'G{current_section_top_row}'].border = thick_top_border

    ws[f'H{current_section_top_row}'].value = f"=SUM(H7:H{len(data_to_write) + 6})"
    ws[f'H{current_section_top_row}'].font = Font(name='Arial Cyr', size=11)
    ws[f'H{current_section_top_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'H{current_section_top_row}'].border = thick_top_right_border
    ws[f'H{current_section_top_row}'].number_format = FORMAT_CURRENCY_RUB

    # Totals Row #2 - Merge cells from D to F and add text
    # Итоговая строка №2 - Объединяем ячейки от D до F и добавляем текст
    current_section_top_row += 1
    ws.merge_cells(f'D{current_section_top_row}:F{current_section_top_row}')
    merged_cell = ws[f'D{current_section_top_row}']
    merged_cell.value = "Сод.в страховании:"
    merged_cell.font = Font(name='Arial Cyr', size=11)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thick_left_border
    ws[f'E{current_section_top_row}'].border = thin_border
    ws[f'F{current_section_top_row}'].border = thin_border
    ws[f'G{current_section_top_row}'].border = thin_border
    ws[f'H{current_section_top_row}'].border = thick_right_border
    ws[f'H{current_section_top_row}'].number_format = FORMAT_CURRENCY_RUB

    # Totals Row #3 - Merge cells from D to F and add text
    # Итоговая строка №3 - Объединяем ячейки от D до F и добавляем текст
    current_section_top_row += 1
    ws.merge_cells(f'D{current_section_top_row}:F{current_section_top_row}')
    merged_cell = ws[f'D{current_section_top_row}']
    merged_cell.value = "Страховка:"
    merged_cell.font = Font(name='Arial Cyr', size=11)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thick_bottom_left_border
    ws[f'E{current_section_top_row}'].border = thick_bottom_border
    ws[f'F{current_section_top_row}'].border = thick_bottom_border
    ws[f'G{current_section_top_row}'].border = thick_bottom_border
    ws[f'H{current_section_top_row}'].border = thick_bottom_right_border
    ws[f'H{current_section_top_row}'].number_format = FORMAT_CURRENCY_RUB

    # save fifth part - NDS row
    # сохраните пятую часть - строка НДС
    current_section_top_row += 6
    ws.merge_cells(f'A{current_section_top_row}:F{current_section_top_row}')
    merged_cell = ws[f'A{current_section_top_row}']
    merged_cell.value = "Тарифы представлены в рублях с учётом НДС."
    merged_cell.font = Font(name='Arial Cyr', size=11)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set the print area to a specific range
    # Установите область печати в определенном диапазоне
    ws.print_area = f"A1:H{current_section_top_row}"

    # Save the workbook
    # Сохраните рабочую книгу
    wb.save(file_path)


def create_excel_files(data, output_dir):
    # Создает новые файлы Excel для каждого уникального идентификатора клиента с помощью pandas.
    """
    Creates new Excel files for each unique client_id using pandas.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Group the data by 'Клиент' and create a DataFrame for each group
    # Сгруппируйте данные по 'Client' и создайте фрейм данных для каждой группы
    clients = data['Клиент'].drop_duplicates()

    # Iterate over each group (DataFrame)
    # Выполнить итерацию по каждой группе (фрейму данных)
    for client in clients:
        # Create a new Excel file for each group
        # Создайте новый файл Excel для каждой группы
        file_name = f"{transform_string(client)}.xlsx"
        file_path = os.path.join(output_dir, file_name)

        # Write the group data to the Excel file
        # Запишите данные группы в файл Excel
        filtered_data = data.loc[data['Клиент'] == client, [
            'Накладная',
            'Дата',
            'Город отправления',
            'Город получения',
            'Вес,кг',
            'Об. вес,кг'
        ]]

        filtered_data = filtered_data.sort_values(by='Дата')
        company_name = data.loc[data['Клиент'] == client, 'Клиент'].values[0]
        company_id = data.loc[data['Клиент'] == client, '№ плател.'].values[0]

        write_to_excel(df=filtered_data, company_name=company_name, company_id=company_id, file_path=file_path)

        # do it once, for testing purposes
        # сделайте это один раз, в целях тестирования
        # break


def zip_excel_files(output_dir, zip_file_name):
    """
    Zips all the new Excel files in the output directory.
    """
    with zipfile.ZipFile(zip_file_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                file_path = os.path.join(root, file)
                zipf.write(file_path, os.path.relpath(file_path, output_dir))


def remove_folder_with_contents(folder_path):
    try:
        # Check if the folder exists
        # Проверьте, существует ли эта папка
        if os.path.exists(folder_path):
            # Remove the folder and all its contents
            # Удалите папку и все ее содержимое
            shutil.rmtree(folder_path)
            print(f"Folder '{folder_path}' and all its contents have been removed.")
        else:
            print(f"The folder '{folder_path}' does not exist.")
    except PermissionError:
        print(f"Permission denied for folder '{folder_path}'.")


def main():
    # Get the directory of the current script
    # Получить каталог текущего скрипта
    script_directory = get_script_folder()
    print(f"{script_directory=}")
    os.chdir(script_directory)

    # Construct the path to the input.xlsx file
    # Построить путь до файла input.xlsx
    input_file_path = os.path.join(script_directory, 'input.xlsx')
    print(f"{input_file_path=}")
    output_dir = 'output'
    zip_file_name = 'output.zip'
    sheet_name = 'Накладные'

    data = read_excel_file(input_file_path, sheet_name)

    create_excel_files(data, output_dir)
    # The rest of the script remains the same for zipping the files
    # Остальная часть скрипта для архивации файлов остается прежней
    zip_excel_files(output_dir, zip_file_name)
    remove_folder_with_contents(output_dir)
    print(f"Excel files zipped successfully: {zip_file_name}")


if __name__ == "__main__":
    main()
