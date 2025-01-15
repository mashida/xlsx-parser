import os
from openpyxl import load_workbook


def process_excel_files():
    # Получаем список всех .xlsx файлов в текущей директории
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]

    for file in excel_files:
        print(f"Обрабатываем файл: {file}")

        # Загружаем workbook
        wb = load_workbook(file)
        sheet = wb.active

        # Начинаем проверку с 7-й строки
        for row in range(7, sheet.max_row + 1):
            cell_C = sheet.cell(row=row, column=3).value
            cell_D = sheet.cell(row=row, column=4).value
            cell_E = sheet.cell(row=row, column=5).value
            cell_F = sheet.cell(row=row, column=6).value

            if (cell_C == 'Череповец' and
                    cell_D == 'Череповец' and
                    isinstance(cell_E, float) and cell_E <= 1 and
                    isinstance(cell_F, float) and cell_F <= 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 200

        # Сохраняем изменения
        wb.save(file)
        print(f"Файл {file} обновлен")


# Вызываем функцию
process_excel_files()
