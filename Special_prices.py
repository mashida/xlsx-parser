"""
Этапы
1. Прочитать файлы xlsx
2. Подготовить данные из ячеек для обработки
3. Применить логику расчета
4. Записать результат в файл
5. Поиск ошибок и улучшения
"""
import os
from openpyxl import load_workbook


def check(input_string):
    if isinstance(input_string, str):
        if ',' in input_string:
            result = input_string.replace(',', '.')
        else:
            result = input_string
    elif isinstance(input_string, int):
        # Если это целое число, просто вернем его как строку
        result = str(input_string)
    else:
        # Для других типов данных вернем исходную строку
        result = str(input_string)
    return result


# 1. Читаем файлы.
def process_excel_files():
    count = 0
    # Получаем список всех .xlsx файлов в директории WB
    target_folder = '/Users/Stanislav_Egorov/Documents/GitHub/SE_Python/xlsx_parser_v2.0/WB'
    excel_files = [f for f in os.listdir(target_folder) if f.endswith('.xlsx')]
    for file in excel_files:
        print(f"Обрабатываем файл: {file}")

        # Загружаем workbook
        file_path = os.path.join(target_folder, file)
        wb = load_workbook(file_path)
        sheet = wb.active
        # 2. Подготовка данных.
        # Начинаем проверку с 7-й строки
        for row in range(7, sheet.max_row + 1):
            cell_c = sheet.cell(row=row, column=3).value
            cell_d = sheet.cell(row=row, column=4).value
            cell_e = float(check(sheet.cell(row=row, column=5).value)) if sheet.cell(row=row,
                                                                                     column=5).value is not None else 1.0
            cell_f = float(check(sheet.cell(row=row, column=6).value)) if sheet.cell(row=row,
                                                                                     column=6).value is not None and sheet.cell(
                row=row, column=7).value != 0.0 else 1.0

            # 3. Логика расчета
            weight: int = 0
            weight_v: int = 0
            if cell_e <= 1.0:
                weight = 1
            elif 1.0 < cell_e <= 2.0:
                weight = 2
            elif 2.0 < cell_e <= 3.0:
                weight = 3
            elif 3.0 < cell_e <= 4.0:
                weight = 4
            elif 4.0 < cell_e <= 5.0:
                weight = 5
            elif cell_e > 5:
                weight = 100000 # жуткий костыль
            if cell_f <= 1.0:
                weight_v = 1
            elif 1.0 < cell_f <= 2.0:
                weight_v = 2
            elif 2.0 < cell_f <= 3.0:
                weight_v = 3
            elif 3.0 < cell_f <= 4.0:
                weight_v = 4
            elif 4.0 < cell_f <= 5.0:
                weight_v = 5
            elif cell_f > 5:
                weight_v = 100000 # жуткий костыль
            max_weight = weight if weight > weight_v else weight_v
            if (cell_c == 'Череповец' and
                    cell_d == 'Череповец' and
                    max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 200
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Череповец' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 220
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Череповец' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 240
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Череповец' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 260
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Череповец' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 280
                count += 1

            elif (cell_c == 'Череповец' and
                  cell_d == 'Санкт-Петербург' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 500
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Санкт-Петербург' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 600
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Санкт-Петербург' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 700
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Санкт-Петербург' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 800
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Санкт-Петербург' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 900
                count += 1

            elif (cell_c == 'Санкт-Петербург' and
                  cell_d == 'Череповец' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 500
                count += 1
            elif (cell_c == 'Санкт-Петербург' and
                  cell_d == 'Череповец' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 600
                count += 1
            elif (cell_c == 'Санкт-Петербург' and
                  cell_d == 'Череповец' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 700
                count += 1
            elif (cell_c == 'Санкт-Петербург' and
                  cell_d == 'Череповец' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 800
                count += 1
            elif (cell_c == 'Санкт-Петербург' and
                  cell_d == 'Череповец' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 900
                count += 1

            elif (cell_c == 'Череповец' and
                  cell_d == 'Москва' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 600
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Москва' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 700
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Москва' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 800
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Москва' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 900
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Москва' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 1000
                count += 1

            elif (cell_c == 'Москва' and
                  cell_d == 'Череповец' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 600
                count += 1
            elif (cell_c == 'Москва' and
                  cell_d == 'Череповец' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 700
                count += 1
            elif (cell_c == 'Москва' and
                  cell_d == 'Череповец' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 800
                count += 1
            elif (cell_c == 'Москва' and
                  cell_d == 'Череповец' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 900
                count += 1
            elif (cell_c == 'Москва' and
                  cell_d == 'Череповец' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 1000
                count += 1

            elif (cell_c == 'Череповец' and
                  cell_d == 'МО' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 600
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'МО' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 700
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'МО' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 800
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'МО' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 900
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'МО' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 1000
                count += 1

            elif (cell_c == 'МО' and
                  cell_d == 'Череповец' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 600
                count += 1
            elif (cell_c == 'МО' and
                  cell_d == 'Череповец' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 700
                count += 1
            elif (cell_c == 'МО' and
                  cell_d == 'Череповец' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 800
                count += 1
            elif (cell_c == 'МО' and
                  cell_d == 'Череповец' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 900
                count += 1
            elif (cell_c == 'МО' and
                  cell_d == 'Череповец' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 1000
                count += 1

            elif (cell_c == 'Череповец' and
                  cell_d == 'Вологда' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 300
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Вологда' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 340
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Вологда' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 380
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Вологда' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 420
                count += 1
            elif (cell_c == 'Череповец' and
                  cell_d == 'Вологда' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 460
                count += 1

            elif (cell_c == 'Вологда' and
                  cell_d == 'Череповец' and
                  max_weight == 1):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 300
                count += 1
            elif (cell_c == 'Вологда' and
                  cell_d == 'Череповец' and
                  max_weight == 2):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 340
                count += 1
            elif (cell_c == 'Вологда' and
                  cell_d == 'Череповец' and
                  max_weight == 3):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 380
                count += 1
            elif (cell_c == 'Вологда' and
                  cell_d == 'Череповец' and
                  max_weight == 4):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 420
                count += 1
            elif (cell_c == 'Вологда' and
                  cell_d == 'Череповец' and
                  max_weight == 5):
                # Вносим изменение
                sheet.cell(row=row, column=8).value = 460
                count += 1
            elif weight == 100000 or weight_v == 100000:
                continue
        print(f"Количество внесенных изменений: {count}")
        # Сохраняем изменения
        wb.save(file_path)
        print(f"Файл {file} обновлен")


def main():
    process_excel_files()


if __name__ == "__main__":
    main()